#!/usr/bin/env python
# coding: utf-8

from openpyxl import load_workbook
import re


class KnudExcel:
    """
    Parameters:
    filename:     path to Excel-Filename
    data_flag:    True:  read values only
                  False: read Excel formulas

    Interface:
    wb:           Excel-workbook
    complete:     dictionary with complete excel-content
    sheets:       list of sheet names
    sheets_upper: list of uppercase sheet names
    get_names():  returns list of cell names and name coordinates
    """

    sheets_upper = {}
    sheets = []

    def __init__(self, filename, data_flag=True):
        self.wb = load_workbook(filename, data_only=data_flag)
        self.complete = self.__get_complete()
        self.LINES = 500

    @staticmethod
    def __mk_int(strng):
        """Gibt den alphanumerischen Teil einer Zellenbezeichnung als Zahl zurück"""
        if len(strng) == 1:
            return ord(strng) - 64
        else:
            return (ord(strng[0]) - 64) * 26 + ord(strng[1]) - 64

    def __get_rowcol(self, val):
        """Aus der Zellenbezeichnung BF17 wird Zeilennummer und Spaltennummer extrahiert"""
        row = int(re.sub("[^0-9]", "", val))
        col = self.__mk_int(re.sub("[^A-Z]", "", val))
        return [row, col]

    def __mk_dict(self, ws):
        """Konvertiert ein Sheet einer Exceltabelle in ein Dictionary"""
        dct = {}
        for row in ws.rows:
            for item in row:
                if item.value is not None:
                    bez = str(item).split(".")[1]
                    bez = bez[:-1]
                    rowcol = self.__get_rowcol(bez)
                    row = rowcol[0]
                    col = rowcol[1]
                    dct[bez] = [row, col, item.value]
        return dct

    def __get_value(self, sheet, cell):
        """ermittelt den Wert einer Zelle cell im Sheet sheet"""
        dct = self.complete[sheet]
        try:
            value = dct[cell][2]
        except Exception:
            value = ""
        return value

    def __get_complete(self):
        """Schreibt die komplette Excel-Datei in ein Dictionary"""
        complete = {}
        sheets_upper = {}
        for sheet in self.wb.sheetnames:
            ws = self.wb[sheet]
            complete[sheet] = self.__mk_dict(ws)
            sheets_upper[sheet.upper()] = sheet
            self.sheets.append(sheet)
        self.sheets_upper = sheets_upper
        print("complete")
        return complete

    @staticmethod
    def __get_cellname(row, col):
        """Berechnet aus Zeilen- und Spaltennummer den Namen der Zelle"""
        return chr(col + 64) + str(row)

    def __mk_fake_call(self, term: str):
        """retrieve entry-fact-assumption constant and generate call
        fake = facts, assumptions, k and entries"""
        try:
            name = term
            term_splitted = re.split(r"[!]", term)
            sector = term_splitted[0]
            cell = term_splitted[1]
            cont = self.complete[sector][cell]
            if sector == "Fakten":
                newcell = "C" + str(cont[0])
                name = self.complete[sector][newcell][2]
                return "fact('" + name + "')"
            elif sector == "Annahmen":
                newcell = "C" + str(cont[0])
                name = self.complete[sector][newcell][2]
                return "ass('" + name + "')"
            elif sector == "Eingabe":
                newcell = "B" + str(cont[0])
                name = self.complete[sector][newcell][2]
                return "entry('" + name + "')"
            else:
                return name
        except Exception as e:
            # print('key-error:', sector, e)
            return None

    def get_names(self):
        """
        Schreibt alle Namen der Excel-Datei in ein Dictionary
        returns names and name_coords as list
        """
        names = {}
        name_coords = {}

        parameter_names = [i.name for i in self.wb.defined_names.definedName]
        for parameter in parameter_names:
            try:
                area = self.wb.defined_names[parameter]
                try:
                    for sheet_title, sheet_coords in area.destinations:
                        coords = sheet_coords.replace("$", "")
                        value = self.__get_value(sheet_title, coords)
                        # name = sheet_title + "!" + parameter
                        name = parameter
                        names[name] = value
                        coords_sheet = sheet_title + "!" + coords
                        name_coords[name] = coords_sheet

                except Exception:
                    # print(parameter, end = " ")
                    print(".", end="")
            except Exception:
                print("*", end="")
        return [names, name_coords]

    def code_gen(self, sector: str, epoch: int, fn: str, print_cell: bool = False):
        """
        Python Code generation for specified sector

        Parameters:
            sector: Sector to choose (first letter lower case)
            epoch:  18 or 30 calculation for 2018 or 2030
            fn:     path to filename for output
            print_cell (False=default, True) prints also excel-cell name of calculated cell

        Example:
        excel = KnudExcel('./03_KStP_Generator_v2a.xlsx', False)
        excel.code_gen('h', 18, 'code_file') # prints heat2018 formulae in 'code_file'
        excel.code_gen('i', 30, 'code_file', True) # prints also Excel cell names
        """

        """Define row ranges of sectors (may change with Excel version"""
        h_range = range(8, 37)
        e_range = range(37, 100)
        f_range = range(100, 123)
        r_range = range(123, 156)
        b_range = range(156, 186)
        i_range = range(186, 231)
        t_range = range(231, 284)
        a_range = range(284, 341)
        l_range = range(341, 388)
        c18_range = range(6, 49)
        c30_range = range(49, 143)

        """Excel Column range for epoch"""
        cols = []
        if epoch == 18:
            cols = c18_range
        if epoch == 30:
            cols = c30_range

        """Select sector"""
        rows = []
        if sector == "h":
            rows = h_range
        if sector == "e":
            rows = e_range
        if sector == "f":
            rows = f_range
        if sector == "r":
            rows = r_range
        if sector == "b":
            rows = b_range
        if sector == "i":
            rows = i_range
        if sector == "t":
            rows = t_range
        if sector == "a":
            rows = a_range
        if sector == "l":
            rows = l_range

        excel_calc = self.complete["Berechnung"]
        """replace excel cells by (row, column) tuple """
        edict = {(v[0], v[1]): v[2] for (k, v) in excel_calc.items()}
        """dictionary of column names"""
        cdict = {t[1]: v for (t, v) in edict.items() if t[0] == 3}
        """dictionary of row names"""
        rdict = {t[0]: v for (t, v) in edict.items() if t[1] == 1}

        names, name_coords = self.get_names()
        try:
            with open(fn, "w") as out:
                for i in rows:
                    if i not in rdict:
                        continue
                    for j in cols:
                        t = (i, j)
                        if t in edict:
                            expr = edict[t]
                            if type(expr) == str:
                                expr = expr.replace("$", "")  # .replace('=','')
                                elist = re.split(
                                    r"[\*, \/, \-, \+, \:, \=, \(, \)]", expr
                                )
                                for item in elist:
                                    if item in excel_calc:
                                        var = self.__get_rowcol(item)
                                        if var[0] not in rdict or var[1] not in cdict:
                                            continue

                                        if var[1] in c18_range:
                                            suffix = "18."
                                        elif var[1] in c30_range:
                                            suffix = "30."

                                        if var[0] in h_range:
                                            cl = "h" + suffix
                                        elif var[0] in e_range:
                                            cl = "e" + suffix
                                        elif var[0] in f_range:
                                            cl = "f" + suffix
                                        elif var[0] in r_range:
                                            cl = "r" + suffix
                                        elif var[0] in b_range:
                                            cl = "b" + suffix
                                        elif var[0] in i_range:
                                            cl = "i" + suffix
                                        elif var[0] in t_range:
                                            cl = "t" + suffix
                                        elif var[0] in a_range:
                                            cl = "a" + suffix
                                        elif var[0] in l_range:
                                            cl = "l" + suffix
                                        # do not write current sector
                                        if cl == sector + str(epoch) + ".":
                                            cl = ""

                                        expr = expr.replace(
                                            item,
                                            cl + rdict[var[0]] + "." + cdict[var[1]],
                                        )
                                    else:
                                        item = item.strip()
                                        m = re.match("[\w]+![\w]+", item)
                                        if m:
                                            replace_item = self.__mk_fake_call(item)
                                            expr = expr.replace(item, replace_item)
                                        else:
                                            if item in name_coords:
                                                replace_item = self.__mk_fake_call(
                                                    name_coords[item]
                                                )
                                                if replace_item is not None:
                                                    expr = expr.replace(
                                                        item, replace_item
                                                    )

                                expr = expr.replace("=", " = (")
                                expr = expr.replace("+", " + ")
                                expr = expr.replace("-", " - ")
                                expr = expr.replace("*", " * ")
                                expr = expr.replace("/", " / ")
                                if j not in cdict:
                                    continue

                                # print formula
                                if print_cell:
                                    cell = chr(int((j - 1) / 26) - 1 + 65)
                                    cell = cell + chr((j - 1) % 26 + 65)
                                    cell = cell.lower() + str(i)
                                    print(
                                        cell
                                        + "\t"
                                        + rdict[i]
                                        + "."
                                        + cdict[j]
                                        + expr
                                        + ")",
                                        file=out,
                                    )
                                else:
                                    print(
                                        "\t" + rdict[i] + "." + cdict[j] + expr + ")",
                                        file=out,
                                    )

        except Exception:
            raise


class SheetCompare:
    """
    class SheetCompare generates two text files from two different versions of an excel sheet
    for use in a file compare program

    Example:
    from knud_classes import *

    path1 = '<path to first excel file>'
    path2 = '<path to second excel file>'

    knud_excel1 = KnudExcel(path1)
    knud_excel2 = KnudExcel(path2)

    excel1 = knud_excel1.complete
    excel2 = knud_excel2.complete

    sc = SheetCompare(excel1, excel2)
    sc.set_sheet('Eingabe', 4)

    sc.write_files('<path to fist output file>', <'path to secnd output file'>)
    """

    """interface"""

    def __init__(self, excel1: dict, excel2: dict):
        """excel1 and excel2 are the complete dictionaries from 2 Excel tables
        generated by the Knud-Class"""
        self.excel1 = excel1
        self.excel2 = excel2
        self.sheet_title = "Annahmen"
        self.cols = 5

    def set_sheet(self, sheet_title: str, number_columns: int):
        """sheet_title: name of the sheet to print
        number_columns: number of columns (from 1st) to consider
        """
        self.sheet_title = sheet_title
        self.cols = number_columns

    def write_files(self, file1: str, file2: str):
        """file1, file2: file paths to the output files"""
        self.__write_file(file1, self.excel1)
        self.__write_file(file2, self.excel2)

    """private methods"""

    def __sheet_content(self, sheet: dict):
        d = {}
        for k, v in sheet.items():
            rownr = v[0]
            row = [None] * self.cols
            # transform dict and make row list
            dict1 = {(v[0], v[1]): v[2] for (k, v) in sheet.items() if v[0] == rownr}
            for kk, vv in dict1.items():
                if self.cols + 1 > kk[1] > 0:
                    row[kk[1] - 1] = vv
            d[rownr] = row
        return d

    def __write_file(self, fn, excel: dict):
        with open(fn, "w") as out:
            try:
                d = self.__sheet_content(excel[self.sheet_title])
                counter = 0
                for key in d:
                    counter = counter + 1
                    # print empty lines
                    while key > counter:
                        print("[]", file=out)
                        counter = counter + 1
                    print(d[key], file=out)

            except Exception as e:
                print("Error: ", e)


def formula_sort(in_file: str, out_file: str):
    """sorts a set of formulas by call
    in_file: input File
    out_file: output File"""
    try:
        """open input file"""
        with open(in_file, "r") as file:
            content = file.read()
        """ prepare file """
        txt = content.replace("\t", "").split("\n")
        if txt[-1] == "":
            txt.remove(txt[-1])

        """sort algo"""
        big_num = 9999
        pivot = 0  # pivot variable index
        swaps = big_num
        varlist = []
        for i in range(0, len(txt)):
            var = re.search("[\w\.]+", txt[i]).group(0)
            varlist.append(var)

        count = 0
        while swaps > 0:
            count = count + 1
            swaps = 0
            for var in varlist:
                dependent = big_num
                for i in range(0, len(txt)):
                    if re.search("(?<![\w\.])" + var + " =", txt[i]):
                        pivot = i
                    elif re.search("(?<![\w\.])" + var + "(?!\w)", txt[i]):
                        if i < dependent:
                            dependent = i

                """ if dependent variable is used before definition swap formula"""
                if dependent < pivot:
                    swaps = swaps + 1
                    txt[pivot], txt[dependent] = txt[dependent], txt[pivot]
            print("iter ", count, "\t", swaps, "swaps")

        with open(out_file, "w") as out:
            for line in txt:
                print(line, file=out)

    except Exception as e:
        print("Error: ", e)
