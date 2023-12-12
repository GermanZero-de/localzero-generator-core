# pyright: strict
from dataclasses import dataclass
import graphlib
import openpyxl
import sys
import re
import typing
import datetime
import csv

PRELUDE = """\"\"\"
This module was auto generated from an annotated version of the 2018 facts file, which
contained explicit formulas for every derived fact. This way we could simplify updating
the facts, without changing a lot of the actual code.
\"\"\"

from . import refdata

def calculate_derived_facts(rd: refdata.RefData):
    import sys
    f = rd.facts()
"""
ROWS: typing.TypeAlias = list[dict[str, str | float | datetime.datetime | None]]

FACT_REGEX = re.compile(r"(Fa[ck]t_[A-Za-z0-9_]+)")


def replace_fact_name_by_fact_lookup(fact_name: str) -> str:
    return f'f.fact("{fact_name}")'


def rows_with_header(w: openpyxl.Workbook) -> ROWS:
    header: tuple[str | float | datetime.datetime | None] | None = None
    sheet = w["2018"]
    for header_row in sheet.iter_rows(
        min_row=1, max_row=1, max_col=11, values_only=True
    ):
        header = header_row  # type: ignore
    if header is None:
        raise Exception("Could not find header row")
    rows: ROWS = []
    for r in sheet.iter_rows(min_row=2, max_col=11, max_row=2000, values_only=True):
        rows.append(
            dict(
                [
                    (h, x)
                    for (h, x) in zip(header, r)
                    if h is not None and type(h) is str
                ]
            )
        )
    return rows


@dataclass
class DerivedFact:
    code: str
    label: str
    depends_on: set[str]
    original_row_num: int


def gen_calculate_derived_facts(rows: ROWS):
    print(PRELUDE)
    derived_facts: dict[str, DerivedFact] = {}
    row_num = 0
    all_derived_facts: set[str] = set()
    for data in rows:
        if data["update 2022"] == "xF":
            if data["Formula"] is None or data["Formula"] == "noch nicht existent":
                continue
            label: str = data["label"]  # type: ignore
            all_derived_facts.add(label)

    for data in rows:
        row_num += 1
        if data["update 2022"] == "xF":
            if data["Formula"] is None or data["Formula"] == "noch nicht existent":
                continue
                # raise Exception(f"Missing formula for {data['label']}")
            label: str = data["label"]  # type: ignore
            formula: str = data["Formula"]  # type: ignore
            formula = FACT_REGEX.sub(
                lambda m: replace_fact_name_by_fact_lookup(m.group(1)), formula
            )
            del data["Formula"]
            del data["label"]
            del data["update 2022"]
            del data["value"]
            data = {k: ("" if v is None else v) for k, v in data.items()}
            code = f"""    f.add_derived_fact("{label}", {formula}, {data})"""
            dependencies = set(FACT_REGEX.findall(formula)) & all_derived_facts
            df = DerivedFact(
                code=code,
                label=label,
                depends_on=dependencies,
                original_row_num=row_num,
            )
            derived_facts[label] = df

    tsorter = graphlib.TopologicalSorter(
        {df.label: df.depends_on for df in derived_facts.values()}
    )
    for df in list(tsorter.static_order()):
        print(derived_facts[df].code)


def list_derived_facts(rows: ROWS):
    for data in rows:
        if data["update 2022"] == "F":
            if data["Formula"] is None or data["Formula"] == "noch nicht existent":
                continue
            print(data["label"])


def extract_new_facts(rows: ROWS):
    columns = [
        "label",
        "group",
        "description",
        "value",
        "unit",
        "rationale",
        "reference",
        "link",
    ]
    with open("new_facts.csv", "w", encoding="utf-8") as fp:
        writer = csv.writer(fp, lineterminator="\n")
        for data in rows:
            if data["update 2022"] in ["xNEW"] and data["value"] is not None:
                row = [data[c] for c in columns]
                row = [d if type(d) != str else d.replace("\n", " ") for d in row]
                writer.writerow(row)


def main():
    mode: typing.Literal[
        "gen_calculate_derived_facts",
        "list_derived_facts",
        "extract_new_facts",
        "usage",
    ]
    filename: str = ""
    match sys.argv:
        case [_, f, "gen_calculate_derived_facts"]:
            filename = f
            mode = "gen_calculate_derived_facts"
        case [_, f, "list_derived_facts"]:
            filename = f
            mode = "list_derived_facts"
        case [_, f, "extract_new_facts"]:
            filename = f
            mode = "extract_new_facts"
        case _:
            print(
                f"Usage: python  <path-to-2018_facts_edit_2022.xlsx> gen_calculate_derived_facts|extract_new_facts"
            )
            sys.exit(1)

    w = openpyxl.load_workbook(filename, data_only=True)
    rows = rows_with_header(w)
    match mode:
        case "gen_calculate_derived_facts":
            gen_calculate_derived_facts(rows)
        case "list_derived_facts":
            list_derived_facts(rows)
        case "extract_new_facts":
            extract_new_facts(rows)


main()
