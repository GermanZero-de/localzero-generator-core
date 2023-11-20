# pyright: strict
from dataclasses import dataclass
import graphlib
import openpyxl
import sys
import re
import typing
import datetime
import csv

PRELUDE = """\"\"\"
This module was auto generated from an annotated version of the 2018 assumptions file, which
contained explicit formulas for every derived assumption. This way we could simplify updating
the assumptions, without changing a lot of the actual code.
\"\"\"

from . import refdata

def calculate_derived_assumptions(rd: refdata.RefData):
    import sys
    a = rd.assumptions()
    f = rd.facts()
"""
ROWS: typing.TypeAlias = list[dict[str, str | float | datetime.datetime | None]]

ASSUMPTION_REGEX = re.compile(r"(Ass_[A-Za-z0-9_]+)")


def replace_ass_name_by_ass_lookup(ass_name: str) -> str:
    return f'a.ass("{ass_name}")'


FACT_REGEX = re.compile(r"(Fa[ck]t_[A-Za-z0-9_]+)")


def replace_fact_name_by_fact_lookup(fact_name: str) -> str:
    return f'f.fact("{fact_name}")'


def rows_with_header(w: openpyxl.Workbook) -> ROWS:
    header: tuple[str | float | datetime.datetime | None] | None = None
    sheet = w["2018_assumptions_edit_2022"]
    for header_row in sheet.iter_rows(
        min_row=1, max_row=1, max_col=11, values_only=True
    ):
        header = header_row  # type: ignore
    if header is None:
        raise Exception("Could not find header row")
    rows: ROWS = []
    for r in sheet.iter_rows(min_row=2, max_col=11, max_row=2000, values_only=True):
        rows.append(
            dict(
                [
                    (h, x)
                    for (h, x) in zip(header, r)
                    if h is not None and type(h) is str
                ]
            )
        )
    return rows


@dataclass
class DerivedAss:
    code: str
    label: str
    depends_on: set[str]
    original_row_num: int


FORMULA = "FORMULA (ONLY WHEN DATA CHANGABLE)"


def gen_calculate_derived_ass(rows: ROWS):
    print(PRELUDE)
    derived_ass: dict[str, DerivedAss] = {}
    row_num = 0
    all_derived_ass: set[str] = set()
    for data in rows:
        if data["update 2022?"] == "F":
            if data[FORMULA] is None or data[FORMULA] == "noch nicht existent":
                continue
            label: str = data["label"]  # type: ignore
            all_derived_ass.add(label)

    for data in rows:
        row_num += 1
        if data["update 2022?"] == "F":
            if data[FORMULA] is None or data[FORMULA] == "noch nicht existent":
                continue
                # raise Exception(f"Missing formula for {data['label']}")
            label: str = data["label"]  # type: ignore
            formula: str = data[FORMULA]  # type: ignore
            formula = ASSUMPTION_REGEX.sub(
                lambda m: replace_ass_name_by_ass_lookup(m.group(1)), formula
            )
            formula = FACT_REGEX.sub(
                lambda m: replace_fact_name_by_fact_lookup(m.group(1)), formula
            )
            del data[FORMULA]
            del data["label"]
            del data["update 2022?"]
            del data["value"]
            data = {k: ("" if v is None else v) for k, v in data.items()}
            code = f"""    a.add_derived_assumption("{label}", {formula}, {data})"""
            dependencies = set(ASSUMPTION_REGEX.findall(formula)) & all_derived_ass
            df = DerivedAss(
                code=code,
                label=label,
                depends_on=dependencies,
                original_row_num=row_num,
            )
            derived_ass[label] = df

    tsorter = graphlib.TopologicalSorter(
        {df.label: df.depends_on for df in derived_ass.values()}
    )
    for df in list(tsorter.static_order()):
        print(derived_ass[df].code)


def list_derived_assumptions(rows: ROWS):
    for data in rows:
        if data["update 2022"] == "F":
            if data["Formula"] is None or data["Formula"] == "noch nicht existent":
                continue
            print(data["label"])


def extract_new_assumptions(rows: ROWS):
    columns = [
        "label",
        "group",
        "description",
        "value",
        "unit",
        "rationale",
        "reference",
        "link",
    ]
    with open("new_assumptions.csv", "w", encoding="utf-8") as fp:
        writer = csv.writer(fp, lineterminator="\n")
        for data in rows:
            if data["update 2022"] == "NEW" and data["value"] is not None:
                row = [data[c] for c in columns]
                row = [d if type(d) != str else d.replace("\n", " ") for d in row]
                writer.writerow(row)


def main():
    mode: typing.Literal[
        "gen_calculate_derived_ass",
        "list_derived_ass",
        "extract_new_ass",
        "usage",
    ]
    filename: str = ""
    match sys.argv:
        case [_, f, "gen_calculate_derived_ass"]:
            filename = f
            mode = "gen_calculate_derived_ass"
        case [_, f, "list_derived_ass"]:
            filename = f
            mode = "list_derived_ass"
        case [_, f, "extract_new_ass"]:
            filename = f
            mode = "extract_new_ass"
        case _:
            print(
                f"Usage: python  <path-to-2018_ass_edit_2022.xlsx> gen_calculate_derived_ass|extract_new_ass"
            )
            sys.exit(1)

    w = openpyxl.load_workbook(filename, data_only=True)
    rows = rows_with_header(w)
    match mode:
        case "gen_calculate_derived_ass":
            gen_calculate_derived_ass(rows)
        case "list_derived_ass":
            list_derived_assumptions(rows)
        case "extract_new_ass":
            extract_new_assumptions(rows)


main()
