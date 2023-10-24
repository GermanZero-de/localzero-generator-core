import enum
import dataclasses
import datetime
import re
import sys


class ChangeKind(enum.Enum):
    DISSOLUTION = enum.auto()
    PARTIAL_SPIN_OFF = enum.auto()
    CHANGE = enum.auto()

    def __str__(self):
        return self.name

    @classmethod
    def of_str(cls, s: str) -> "ChangeKind":
        return cls[s]


def change_kind_from_str(s: str) -> ChangeKind:
    """Raise ValueError if the string is not a valid change kind"""
    if s == "1":
        return ChangeKind.DISSOLUTION
    elif s == "2":
        return ChangeKind.PARTIAL_SPIN_OFF
    elif s == "3":
        return ChangeKind.CHANGE
    elif s == "4":
        return ChangeKind.CHANGE
    elif s == "3, 4":
        return ChangeKind.CHANGE
    elif s == "3.4":
        # Sigh looks like this file is manually maintained
        return ChangeKind.CHANGE
    else:
        raise ValueError(f"Invalid change kind: {s}")


@dataclasses.dataclass
class RawRecord:
    change_kind: ChangeKind
    ags: str
    name: str
    effective_date: datetime.date
    new_ags: str
    new_name: str
    area_in_sqm: int | None
    population: int | None

    def __str__(self):
        return f"{self.effective_date}: {self.change_kind} {self.ags} ({self.name}) -> {self.new_ags} ({self.new_name})"


GERMAN_DATE_REGEX = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")


def must_be_str(s: object, name: str) -> str:
    if not isinstance(s, str):
        raise ValueError(f"{name} must be a string, got {s!r}")
    return s


def must_be_int(i: object, name: str) -> int:
    if isinstance(i, str):
        return int(i)
    if not isinstance(i, int):
        raise ValueError(f"{name} must be an int, got {i!r}")
    return i


def must_be_date(d: object, name: str) -> datetime.date:
    if not isinstance(d, datetime.date):
        raise ValueError(f"{name} must be a date, got {d!r}")
    return d


def date_from_german(s: str) -> datetime.date:
    """Parse a date in the format DD.MM.YYYY
    Raises ValueError if the date is invalid.
    """
    m = GERMAN_DATE_REGEX.fullmatch(s)
    if not m:
        raise ValueError(f"Invalid date: {s}")
    day = int(m[1])
    month = int(m[2])
    year = int(m[3])
    return datetime.date(year, month, day)


def load() -> list[RawRecord]:
    """Convert the DeStatis Gebietsänderungen xlsx files, into a list of RawRecord"""

    #     Notes on the files
    #
    #     There are 4 kinds of changes in the file, effective_on is always the date of the change.
    #
    #     1 - Dissolution -- The ags is no longer valid.  That is the relevant commune no longer
    #             exists as an individual entity.
    #         new_ags is the ags of the commune that incorporated the area of the old ags
    #     2 - Partial spin off -- The commune continues to be valid, but some of its area
    #             no longer belongs to it, but another commune.
    #         ags is the ags of the commune that lost the area
    #         new_ags is the ags of the commune that got the area
    #
    #         Partial spin offs always come in multiple entries, one entry for each spinned of area
    #         and one for the remaining area.  Or with other words, the sum of the areas
    #         of all the entries is the area of the commune before the spin off.  And also the
    #         last row of a block of spin offs for the same ags is the one that has the
    #         remaining area.
    #     3,4 - Change of either AGS or name.  The commune and its area hasn't changed, but
    #         the ags or the name has.  In the file this is supposed to be recorded as
    #
    #         3 only AGS change
    #         4 only name change
    #         3,4 AGS and name change
    #
    #         but in practice one can see records where a 3 is both AGS and name change
    #         as well as "3, 4" and "3.4" (SIGH).  So I'm going to treat all of these as
    #         the same kind of change. And one can use the old and new value of the fields
    #         to figure out what changed.
    import openpyxl

    files = [
        ("destatis_2020_Gebietsänderungen_2019.xlsx", "Gebietsaenderungen 2019", 8),
        ("destatis_2021_Gebietsänderungen_2020.xlsx", "Gebietsaenderungen 2020", 8),
        ("destatis_2021_Gebietsänderungen_2021.xlsx", "Gebietsaenderungen 2021", 5),
    ]
    data: list[RawRecord] = []
    for file, sheet_name, first_row in files:
        print(f"Converting sheet {sheet_name} from {file}", file=sys.stderr)
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(
            min_row=first_row,
            max_row=sheet.max_row,
            min_col=1,
            max_col=13,
            values_only=True,
        ):
            # Skip empty rows (or rather rows without a KennZiffer)
            if row[0] is None:
                continue
            # Similar skip entries without a AGS (such as Gemeineverbände)
            if row[3] is None:
                continue
            ags = must_be_str(row[3], "ags")
            name = must_be_str(row[4], "name")
            change_kind = change_kind_from_str(
                str(row[5])
            )  # can be str or int in the sheet
            area_in_sqm = must_be_int(row[6], "area_in_sqm") if row[6] else None
            population = must_be_int(row[7], "population") if row[7] else None
            new_ags = must_be_str(row[9], "new_ags")
            new_name = must_be_str(row[10], "new_name")
            # Can you believe it? We get a spreadsheet but the dates are formatted as strings...
            effective_date = date_from_german(must_be_str(row[11], "effective_date"))

            record = RawRecord(
                change_kind,
                ags,
                name=name,
                effective_date=effective_date,
                new_ags=new_ags,
                new_name=new_name,
                area_in_sqm=area_in_sqm,
                population=population,
            )
            data.append(record)

    return data
